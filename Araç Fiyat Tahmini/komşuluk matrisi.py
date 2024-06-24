import pandas as pd
from sklearn.metrics.pairwise import euclidean_distances
from sklearn.preprocessing import OneHotEncoder
import openpyxl
from openpyxl.styles import PatternFill

# Excel dosyasını oku
df = pd.read_excel('output.xlsx')

# Renk verisini one-hot encoding ile dönüştür
encoder = OneHotEncoder()
renk_encoded = encoder.fit_transform(df[['Renk']]).toarray()

# One-hot encoded renk sütunlarını DataFrame'e ekle
renk_columns = encoder.get_feature_names_out(['Renk'])
df_encoded = pd.concat([df.drop('Renk', axis=1), pd.DataFrame(renk_encoded, columns=renk_columns)], axis=1)

# Komşuluk matrisi hesapla (Yıl, Kilometre, Fiyat ve Renk sütunları kullanarak)
X = df_encoded[['Yıl', 'Kilometre', 'Fiyat'] + list(renk_columns)]
dist_matrix = euclidean_distances(X, X)

# Komşuluk matrisini DataFrame olarak sakla ve Excel dosyasına kaydet
dist_df = pd.DataFrame(dist_matrix, index=df.index, columns=df.index)
dist_df.to_excel('komsuluk_matrisi.xlsx')

# Sıfırdan farklı en küçük değeri bul
min_nonzero_value = dist_df[dist_df > 0].min().min()

# Excel dosyasını yeniden aç ve en küçük değeri kırmızı renge boya
wb = openpyxl.load_workbook('komsuluk_matrisi.xlsx')
ws = wb.active

# Boyama rengi (kırmızı)
fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Sıfırdan farklı en küçük değeri bul ve boya
for row in range(2, ws.max_row + 1):
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value > 0 and cell.value == min_nonzero_value:
            cell.fill = fill

# Değişiklikleri kaydet
wb.save('komsuluk_matrisi.xlsx')
wb.close()
